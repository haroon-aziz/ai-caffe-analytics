from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd

from analytics import aggregates
from config import Config, get_config
from utils.logger import get_logger

logger = get_logger(__name__)

_C = get_config().dashboard
_PRIMARY, _ACCENT, _SUCCESS = _C.primary_color, _C.accent_color, _C.success_color


@dataclass
class ReportSummary:
    generated_at: datetime
    period_start: Optional[datetime]
    period_end: Optional[datetime]
    total_visitors: int
    peak_occupancy: int
    avg_stay_seconds: float
    peak_hour: Optional[int]
    max_queue_length: int
    total_alerts: int
    scope: str

    def as_rows(self) -> List[Tuple[str, str]]:
        peak = f"{self.peak_hour:02d}:00" if self.peak_hour is not None else "—"
        span = "—"
        if self.period_start and self.period_end:
            span = f"{self.period_start:%Y-%m-%d %H:%M} → {self.period_end:%Y-%m-%d %H:%M}"
        return [
            ("Scope", self.scope),
            ("Report period", span),
            ("Total visitors", str(self.total_visitors)),
            ("Peak occupancy", str(self.peak_occupancy)),
            ("Average stay", f"{self.avg_stay_seconds:.0f} s"),
            ("Peak hour", peak),
            ("Max queue length", str(self.max_queue_length)),
            ("Total alerts", str(self.total_alerts)),
            ("Generated", f"{self.generated_at:%Y-%m-%d %H:%M:%S}"),
        ]


class ReportGenerator:
    def __init__(
        self,
        db: Any,
        config: Optional[Config] = None,
        session_id: Optional[str] = None,
    ) -> None:
        self.db = db
        self.config = config or get_config()
        self.session_id = session_id
        self._data = self._gather()


    def _gather(self) -> Dict[str, Any]:
        visits = self.db.get_visits(session_id=self.session_id)
        snaps = self.db.get_snapshots(session_id=self.session_id)
        alerts = self.db.get_alerts(limit=1000)

        visits_df = pd.DataFrame(
            [
                {
                    "track_id": v.track_id,
                    "role": v.role,
                    "entry_time": v.entry_time,
                    "exit_time": v.exit_time,
                    "duration_seconds": v.duration_seconds,
                }
                for v in visits
            ]
        )
        snaps_df = pd.DataFrame(
            [
                {
                    "timestamp": s.timestamp,
                    "current_customers": s.current_customers,
                    "current_staff": s.current_staff,
                    "occupied_tables": s.occupied_tables,
                    "queue_length": s.queue_length,
                    "avg_wait_seconds": s.avg_wait_seconds,
                    "fps": s.fps,
                }
                for s in snaps
            ]
        )
        hourly = aggregates.hourly_visitors(self.db, role="customer")
        daily = aggregates.daily_visitors(self.db, days=30, role="customer")
        stay = aggregates.stay_duration_stats(self.db, role="customer")

        customer_visits = [v for v in visits if v.role == "customer"]
        durations = [v.duration_seconds for v in customer_visits if v.duration_seconds]
        entries = [v.entry_time for v in visits if v.entry_time]
        summary = ReportSummary(
            generated_at=datetime.now(),
            period_start=min(entries) if entries else None,
            period_end=max(entries) if entries else None,
            total_visitors=len({v.track_id for v in customer_visits}),
            peak_occupancy=max((s.current_customers for s in snaps), default=0),
            avg_stay_seconds=(sum(durations) / len(durations)) if durations else 0.0,
            peak_hour=aggregates.peak_hour(self.db, role="customer"),
            max_queue_length=max((s.queue_length for s in snaps), default=0),
            total_alerts=len(alerts),
            scope="Session" if self.session_id else "All history",
        )
        return {
            "visits": visits_df,
            "snapshots": snaps_df,
            "hourly": hourly,
            "daily": daily,
            "stay": stay,
            "alerts": alerts,
            "summary": summary,
        }

    @property
    def summary(self) -> ReportSummary:
        return self._data["summary"]


    def build_csv(self, dataset: str = "visits") -> bytes:
        df: pd.DataFrame = self._data.get(dataset, pd.DataFrame())
        return df.to_csv(index=False).encode("utf-8")


    def build_excel(self, path: Optional[Path] = None) -> bytes:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        buffer = io.BytesIO()
        summary_df = pd.DataFrame(self.summary.as_rows(), columns=["Metric", "Value"])
        hourly_df = pd.DataFrame(
            {"hour": list(self._data["hourly"].keys()),
             "visitors": list(self._data["hourly"].values())}
        )
        daily_df = pd.DataFrame(self._data["daily"], columns=["date", "visitors"])
        stay = self._data["stay"]
        stay_df = pd.DataFrame(
            {"bucket": list(stay.histogram.keys()), "count": list(stay.histogram.values())}
        ) if stay.histogram else pd.DataFrame({"bucket": [], "count": []})
        alerts_df = pd.DataFrame(
            [{"type": a.alert_type, "severity": a.severity, "message": a.message,
              "timestamp": a.timestamp} for a in self._data["alerts"]]
        )

        sheets = {
            "Summary": summary_df,
            "Visits": self._data["visits"],
            "Snapshots": self._data["snapshots"],
            "Hourly": hourly_df,
            "Daily": daily_df,
            "Stay Distribution": stay_df,
            "Alerts": alerts_df,
        }

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            for name, df in sheets.items():
                (df if not df.empty else pd.DataFrame({"(no data)": []})).to_excel(
                    writer, sheet_name=name, index=False
                )

            header_fill = PatternFill("solid", fgColor=_PRIMARY.lstrip("#"))
            header_font = Font(bold=True, color="FFFFFF")
            for name in sheets:
                ws = writer.sheets[name]
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                for col_idx, col in enumerate(ws.columns, start=1):
                    width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(40, width + 3)
                ws.freeze_panes = "A2"

        data = buffer.getvalue()
        if path is not None:
            Path(path).write_bytes(data)
        return data


    def _chart_png(self, kind: str) -> Optional[bytes]:
        fig, ax = plt.subplots(figsize=(6.2, 2.6), dpi=130)
        fig.patch.set_facecolor("white")
        ok = True
        if kind == "hourly":
            data = self._data["hourly"]
            if not any(data.values()):
                ok = False
            ax.bar([f"{h:02d}" for h in data], list(data.values()), color=_PRIMARY)
            ax.set_title("Visitors by Hour")
            ax.set_xlabel("Hour"); ax.set_ylabel("Visitors")
            ax.tick_params(axis="x", labelsize=6, rotation=0)
        elif kind == "occupancy":
            df = self._data["snapshots"]
            if df.empty:
                ok = False
            else:
                ax.plot(df["timestamp"], df["current_customers"], color=_ACCENT, linewidth=2)
                ax.fill_between(df["timestamp"], df["current_customers"], color=_ACCENT, alpha=0.15)
            ax.set_title("Occupancy Trend"); ax.set_ylabel("Customers")
            ax.tick_params(axis="x", labelsize=6, rotation=30)
        elif kind == "stay":
            stay = self._data["stay"]
            if not stay.histogram or stay.count == 0:
                ok = False
            else:
                ax.bar(list(stay.histogram.keys()), list(stay.histogram.values()), color=_SUCCESS)
            ax.set_title("Stay Duration Distribution")
            ax.set_xlabel("Duration"); ax.set_ylabel("Visits")
            ax.tick_params(axis="x", labelsize=7)
        ax.grid(True, alpha=0.2)
        fig.tight_layout()

        if not ok:
            plt.close(fig)
            return None
        out = io.BytesIO()
        fig.savefig(out, format="png", facecolor="white")
        plt.close(fig)
        return out.getvalue()

    def build_pdf(self, path: Optional[Path] = None) -> bytes:
        from fpdf import FPDF

        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        primary = _hex_rgb(_PRIMARY)

        pdf.set_fill_color(*primary)
        pdf.rect(0, 0, 210, 24, style="F")
        pdf.set_xy(10, 7)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 18)
        pdf.cell(0, 10, _pdf_safe("CafeAnalytics — Analytics Report"), ln=1)
        pdf.set_text_color(30, 30, 30)
        pdf.ln(10)

        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 8, "Summary", ln=1)
        pdf.set_font("Helvetica", "", 11)
        for label, value in self.summary.as_rows():
            pdf.set_fill_color(240, 241, 245)
            pdf.cell(70, 8, _pdf_safe(f"  {label}"), border=0, fill=True)
            pdf.cell(0, 8, _pdf_safe(f"  {value}"), border=0, ln=1)
        pdf.ln(4)

        for kind, title in (("hourly", "Visitors by Hour"),
                            ("occupancy", "Occupancy Trend"),
                            ("stay", "Stay Duration Distribution")):
            png = self._chart_png(kind)
            if png is None:
                continue
            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 8, _pdf_safe(title), ln=1)
            pdf.image(io.BytesIO(png), w=180)
            pdf.ln(3)

        pdf.set_y(-12)
        pdf.set_font("Helvetica", "I", 8)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 6, _pdf_safe(f"Generated by CafeAnalytics · {datetime.now():%Y-%m-%d %H:%M}"),
                 align="C")

        data = bytes(pdf.output())
        if path is not None:
            Path(path).write_bytes(data)
        return data


    def generate(self, fmt: str = "pdf", title: Optional[str] = None) -> Path:
        fmt = fmt.lower()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = self.config.paths.reports
        out_dir.mkdir(parents=True, exist_ok=True)

        if fmt == "excel":
            path = out_dir / f"report_{stamp}.xlsx"
            self.build_excel(path)
        elif fmt == "pdf":
            path = out_dir / f"report_{stamp}.pdf"
            self.build_pdf(path)
        elif fmt == "csv":
            path = out_dir / f"visits_{stamp}.csv"
            path.write_bytes(self.build_csv("visits"))
        else:
            raise ValueError(f"Unsupported report format: {fmt}")

        if self.db is not None and hasattr(self.db, "add_report"):
            self.db.add_report(
                title or f"Analytics Report ({self.summary.scope})",
                fmt,
                str(path),
                period_start=self.summary.period_start,
                period_end=self.summary.period_end,
            )
        logger.info("Generated %s report: %s", fmt, path)
        return path


def _hex_rgb(hex_color: str) -> Tuple[int, int, int]:
    h = hex_color.lstrip("#")
    return tuple(int(h[i : i + 2], 16) for i in (0, 2, 4))


_UNICODE_MAP = {
    "—": "-",
    "–": "-",
    "→": "->",
    "·": "-",
    "…": "...",
    "‘": "'", "’": "'", "“": '"', "”": '"',
}


def _pdf_safe(text: str) -> str:
    for uni, repl in _UNICODE_MAP.items():
        text = text.replace(uni, repl)
    return text.encode("latin-1", "replace").decode("latin-1")
